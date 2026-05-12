"""
filtrar_estaciones_A2.py
========================
Filtra el shapefile de estaciones de aforo (IMD 2024) conservando
únicamente las estaciones de la A-2 Madrid–Barcelona que aparecen
en el Excel de referencia (250717_Tráfico_tramos_RCE_2024.xlsx).

Criterio de filtrado:
  El campo CODESTA del shapefile debe coincidir con algún CODESTA de
  la hoja "Estaciones2024" del Excel que tenga CARRETERA == 'A-2'.
  Esto garantiza que la fuente de verdad es el propio dataset oficial,
  sin depender de heurísticos geométricos externos.

Salida:
  - estaciones_A2_filtradas.csv   → tabla con todos los atributos + X/Y
  - estaciones_A2_filtradas.shp   → nuevo shapefile listo para SIG

Requisitos (solo librería estándar + openpyxl):
  pip install openpyxl

Uso:
  python filtrar_estaciones_A2.py
  python filtrar_estaciones_A2.py --shp ruta/archivo.shp \
                                   --xlsx ruta/excel.xlsx \
                                   --output mi_salida
"""

import os
import csv
import struct
import argparse

# ---------------------------------------------------------------------------
# Parámetros por defecto
# ---------------------------------------------------------------------------
DEFAULT_SHP    = "250717_Estaciones2024_IMD.shp"
DEFAULT_XLSX   = "250717_Tráfico_tramos_RCE_2024.xlsx"
DEFAULT_OUTPUT = "estaciones_A2_filtradas"
SHEET_NAME     = "Estaciones2024"
COL_CODESTA    = "CODESTA"
COL_CARRETERA  = "CARRETERA"
TARGET_ROAD    = "A-2"


# ---------------------------------------------------------------------------
# Lectura del Excel con openpyxl
# ---------------------------------------------------------------------------

def leer_codestas_a2_xlsx(xlsx_path, sheet_name, col_codesta, col_carretera, road):
    """
    Lee el Excel y devuelve el conjunto de CODESTAs (str) cuya columna
    CARRETERA sea igual a `road`. Solo se aceptan valores numéricos
    (se descartan UUIDs, vacíos u otros formatos inesperados).
    """
    try:
        import openpyxl
    except ImportError:
        raise SystemExit(
            "ERROR: openpyxl no está instalado.\n"
            "       Ejecuta:  pip install openpyxl"
        )

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise SystemExit(
            f"ERROR: La hoja '{sheet_name}' no existe en {xlsx_path}.\n"
            f"       Hojas disponibles: {wb.sheetnames}"
        )

    ws   = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)

    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    try:
        idx_cod = header.index(col_codesta)
        idx_car = header.index(col_carretera)
    except ValueError as e:
        raise SystemExit(
            f"ERROR: Columna no encontrada en '{sheet_name}': {e}\n"
            f"       Columnas disponibles: {header}"
        )

    codestas = set()
    for row in rows:
        carretera = str(row[idx_car]).strip() if row[idx_car] is not None else ""
        if carretera != road:
            continue
        cod = str(row[idx_cod]).strip() if row[idx_cod] is not None else ""
        if cod and cod.isdigit():          # descartar UUIDs y vacíos
            codestas.add(cod)

    wb.close()
    return codestas


# ---------------------------------------------------------------------------
# Lectura de .dbf
# ---------------------------------------------------------------------------

def read_dbf(path):
    with open(path, "rb") as f:
        header      = f.read(32)
        num_records = struct.unpack("<I", header[4:8])[0]
        header_size = struct.unpack("<H", header[8:10])[0]
        record_size = struct.unpack("<H", header[10:12])[0]

        fields = []
        f.seek(32)
        while True:
            fd = f.read(32)
            if not fd or fd[0] == 0x0D:
                break
            name   = fd[:11].replace(b"\x00", b"").decode("latin-1").strip()
            ftype  = chr(fd[11])
            length = fd[16]
            fields.append((name, ftype, length))

        f.seek(header_size)
        records = []
        for _ in range(num_records):
            row_data = f.read(record_size)
            if not row_data or row_data[0] == 0x2A:
                continue
            row, pos = {}, 1
            for name, ftype, length in fields:
                val = row_data[pos : pos + length].decode("latin-1").strip()
                row[name] = val
                pos += length
            records.append(row)

    return fields, records


# ---------------------------------------------------------------------------
# Lectura de .shp (puntos, shape type 1)
# ---------------------------------------------------------------------------

def read_shp_points(path):
    points = []
    with open(path, "rb") as f:
        f.seek(100)
        idx = 0
        while True:
            rec_header = f.read(8)
            if len(rec_header) < 8:
                break
            content_len = struct.unpack(">I", rec_header[4:8])[0]
            content     = f.read(content_len * 2)
            if len(content) < 4:
                break
            shape_type = struct.unpack("<I", content[:4])[0]
            if shape_type == 1 and len(content) >= 20:
                x = struct.unpack("<d", content[4:12])[0]
                y = struct.unpack("<d", content[12:20])[0]
                points.append((idx, x, y))
            else:
                points.append((idx, None, None))
            idx += 1
    return points


# ---------------------------------------------------------------------------
# Escritura de CSV
# ---------------------------------------------------------------------------

def write_csv(path, fields, records, coords):
    field_names = [f[0] for f in fields] + ["UTM_X", "UTM_Y"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=field_names)
        writer.writeheader()
        for rec, (_, x, y) in zip(records, coords):
            row = dict(rec)
            row["UTM_X"] = f"{x:.2f}" if x is not None else ""
            row["UTM_Y"] = f"{y:.2f}" if y is not None else ""
            writer.writerow(row)
    print(f"  CSV guardado: {path}")


# ---------------------------------------------------------------------------
# Escritura de shapefile (puntos)
# ---------------------------------------------------------------------------

def write_shp(base_path, fields, records, coords, prj_content=None):
    shp_path = base_path + ".shp"
    shx_path = base_path + ".shx"
    dbf_path = base_path + ".dbf"
    prj_path = base_path + ".prj"

    n                = len(records)
    rec_content_len  = 20
    rec_total_bytes  = 8 + rec_content_len
    file_len_words   = (100 + n * rec_total_bytes) // 2

    xs = [c[1] for c in coords if c[1] is not None]
    ys = [c[2] for c in coords if c[2] is not None]
    x_min, x_max = (min(xs), max(xs)) if xs else (0, 0)
    y_min, y_max = (min(ys), max(ys)) if ys else (0, 0)

    def file_header(file_len_w):
        h  = struct.pack(">IIIIII", 9994, 0, 0, 0, 0, 0)
        h += struct.pack(">I", file_len_w)
        h += struct.pack("<II", 1000, 1)
        h += struct.pack("<dddddd", x_min, y_min, x_max, y_max, 0.0, 0.0)
        h += struct.pack("<dd", 0.0, 0.0)
        return h

    shx_len_words = (100 + n * 8) // 2

    with open(shp_path, "wb") as shp, open(shx_path, "wb") as shx:
        shp.write(file_header(file_len_words))
        shx.write(file_header(shx_len_words))
        offset = 50
        for i, (_, x, y) in enumerate(coords):
            clen = rec_content_len // 2
            shp.write(struct.pack(">II", i + 1, clen))
            shp.write(struct.pack("<I", 1))
            shp.write(struct.pack("<dd", x, y))
            shx.write(struct.pack(">II", offset, clen))
            offset += 4 + clen

    rec_size        = 1 + sum(f[2] for f in fields)
    header_size_dbf = 32 + 32 * len(fields) + 1

    with open(dbf_path, "wb") as dbf:
        dbf.write(struct.pack("BBBB", 3, 124, 5, 7))
        dbf.write(struct.pack("<I", n))
        dbf.write(struct.pack("<HH", header_size_dbf, rec_size))
        dbf.write(b"\x00" * 20)
        for fname, ftype, flength in fields:
            name_bytes = fname.encode("latin-1")[:10].ljust(11, b"\x00")
            dbf.write(name_bytes)
            dbf.write(ftype.encode("latin-1"))
            dbf.write(b"\x00" * 4)
            dbf.write(struct.pack("BB", flength, 0))
            dbf.write(b"\x00" * 14)
        dbf.write(b"\x0D")
        for rec in records:
            dbf.write(b" ")
            for fname, ftype, flength in fields:
                val     = rec.get(fname, "")
                encoded = val.encode("latin-1")[:flength]
                dbf.write(encoded.rjust(flength) if ftype in ("N", "F")
                          else encoded.ljust(flength))
        dbf.write(b"\x1A")

    if prj_content:
        with open(prj_path, "w") as prj:
            prj.write(prj_content)

    print(f"  Shapefile guardado: {shp_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Filtra estaciones A-2 del shapefile usando el Excel oficial como referencia"
    )
    parser.add_argument("--shp",    default=DEFAULT_SHP,    help="Ruta al .shp de entrada")
    parser.add_argument("--xlsx",   default=DEFAULT_XLSX,   help="Ruta al Excel de referencia")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Prefijo de los ficheros de salida")
    args = parser.parse_args()

    dbf_in  = args.shp.replace(".shp", ".dbf")
    prj_in  = args.shp.replace(".shp", ".prj")
    out_csv = args.output + ".csv"
    out_shp = args.output

    print(f"\n{'='*62}")
    print(f"  Filtrado de estaciones A-2 Madrid-Barcelona")
    print(f"{'='*62}")
    print(f"  Shapefile entrada  : {args.shp}")
    print(f"  Excel referencia   : {args.xlsx}")
    print(f"  Hoja               : {SHEET_NAME}")
    print(f"  Criterio           : CODESTA en Excel con CARRETERA='{TARGET_ROAD}'")

    # 1. CODESTAs del Excel
    print(f"\n[1/4] Leyendo CODESTAs A-2 del Excel ...")
    codestas_a2 = leer_codestas_a2_xlsx(
        args.xlsx, SHEET_NAME, COL_CODESTA, COL_CARRETERA, TARGET_ROAD
    )
    print(f"      {len(codestas_a2)} CODESTAs A-2 encontrados")
    print(f"      Lista: {sorted(codestas_a2, key=int)}")

    # 2. Leer .prj
    prj_content = None
    if os.path.exists(prj_in):
        with open(prj_in) as pf:
            prj_content = pf.read()

    # 3. Leer shapefile
    print(f"\n[2/4] Leyendo shapefile ...")
    fields, records = read_dbf(dbf_in)
    all_points      = read_shp_points(args.shp)
    print(f"      {len(records)} registros | campos: {[f[0] for f in fields]}")

    if "CODESTA" not in [f[0] for f in fields]:
        raise SystemExit("ERROR: El campo 'CODESTA' no existe en el shapefile.")

    # 4. Filtrar
    print(f"\n[3/4] Aplicando filtro por CODESTA ...")
    mask         = [rec["CODESTA"].strip() in codestas_a2 for rec in records]
    filt_records = [r for r, ok in zip(records, mask) if ok]
    filt_coords  = [c for c, ok in zip(all_points, mask) if ok]
    n_result     = len(filt_records)

    # Avisar si algún CODESTA del Excel no está en el shapefile
    shp_codestas = {rec["CODESTA"].strip() for rec in records}
    sin_match    = codestas_a2 - shp_codestas
    if sin_match:
        print(f"      AVISO: {len(sin_match)} CODESTAs del Excel no aparecen en el shapefile:")
        for c in sorted(sin_match, key=int):
            print(f"        CODESTA {c}")

    print(f"      -> Estaciones resultantes: {n_result}")

    # 5. Guardar
    print(f"\n[4/4] Guardando resultados ...")
    write_csv(out_csv, fields, filt_records, filt_coords)
    write_shp(out_shp, fields, filt_records, filt_coords, prj_content)

    # 6. Resumen
    pks = []
    for rec in filt_records:
        try:
            pks.append(float(rec["PK"]))
        except (ValueError, KeyError):
            pass
    pks.sort()

    prov_names = {
        "28": "Madrid", "19": "Guadalajara", "50": "Zaragoza",
        "22": "Huesca",  "25": "Lleida",      "8":  "Barcelona",
        "08": "Barcelona", "17": "Girona",    "42": "Soria"
    }
    provs = {}
    for rec in filt_records:
        p = rec.get("ID_PROVINC", "?").strip()
        provs[p] = provs.get(p, 0) + 1

    print(f"\n{'='*62}")
    print(f"  RESUMEN")
    print(f"{'='*62}")
    print(f"  Estaciones totales en el shapefile   : {len(records)}")
    print(f"  CODESTAs A-2 en el Excel             : {len(codestas_a2)}")
    print(f"  Estaciones resultantes (interseccion): {n_result}")
    if pks:
        print(f"  PK minimo                            : {pks[0]:.1f} km")
        print(f"  PK maximo                            : {pks[-1]:.1f} km")
        print(f"  Longitud del corredor cubierto       : {pks[-1]-pks[0]:.1f} km")
    print(f"\n  Distribucion por provincia:")
    for prov_id, count in sorted(provs.items()):
        name = prov_names.get(prov_id, f"Prov.{prov_id}")
        print(f"    {name:15s} (ID {prov_id:>3}): {count} estaciones")
    print(f"\n  Archivos generados:")
    print(f"    {out_csv}")
    print(f"    {out_shp}.shp / .dbf / .shx / .prj")
    print(f"{'='*62}\n")


if __name__ == "__main__":
    main()